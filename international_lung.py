# author: 方丈
# use for international-fee

import pandas as pd
from openpyxl import Workbook
import glob
import os

def get_good_name_index(file_name_index):
    goods_name = []
    for i in range(len(file_name_index)):
        split_list = os.path.splitext(file_name_index[i])
        file_name = split_list[-2]
        goods_name.append(file_name)
    return goods_name




flag = 1
per_g = float(input('请输入国际均价: '))
file_name_index = glob.glob(r'*.xlsx')
goods_name = get_good_name_index(file_name_index)
print(file_name_index,goods_name)
people = {}
#people = {cn:{'xxx:[数量,重量]'，'xxx:[数量,价格]','total: 重量*均价+价格'}}




for i,file_path in enumerate(file_name_index):
    print(goods_name[i])
    object_name = goods_name[i]
    choose_mode = float(input('请输入模式(按重量0,按均价1): '))
    if choose_mode == 0:
        per_weight = float(input('请输入均重:'))
        df = pd.read_excel(file_name_index[i],header = None)
        nrows = df.shape[0]
        ncols = df.columns.size
        for i in range(ncols):
            for j in range(2,nrows):
                if pd.notnull(df.iloc[j,i]):
                    people_name = df.iloc[j,i]
                    if people_name not in people:
                        people[people_name] = {'total':0}
                        people[people_name][object_name] = [1,per_weight]
                        #people[people_name]['total'] = people[people_name]['total'] + per_weight*per_g
                    else:
                        if object_name not in people[people_name]:
                            people[people_name][object_name] = [1,per_weight]
                        else:
                            people[people_name][object_name][0] +=1
                            people[people_name][object_name][1] = people[people_name][object_name][1] + per_weight   
                    people[people_name]['total'] = people[people_name]['total'] + per_weight*per_g
                    
                else:
                    break

    if choose_mode == 1:
        per_price = float(input('请输入均价: '))
        df = pd.read_excel(file_name_index[i],header = None)
        nrows = df.shape[0]
        ncols = df.columns.size
        for i in range(ncols):
            for j in range(2,nrows):
                if pd.notnull(df.iloc[j,i]):
                    people_name = df.iloc[j,i]
                    if people_name not in people:
                        people[people_name] = {'total':0}
                        people[people_name][object_name] = [1,per_price]
                    else:
                        if object_name not in people[people_name]:
                            people[people_name][object_name] = [1,per_price]
                        else:
                            people[people_name][object_name][0] +=1
                            people[people_name][object_name][1] = people[people_name][object_name][1]+per_price
                    people[people_name]['total'] = people[people_name]['total'] + per_price
                else:
                    break

wb = Workbook()
ws = wb.active

for x in range(1,len(people)):
    for y in range(1,len(file_name_index)*2):
        ws.cell(row = x,column = y)

row_n = 2
ws.cell(row = 1,column = 1,value = 'cn')
print(object_name)
for i,object_n in enumerate (goods_name):
    ws.cell(row = 1,column = (i+1)*2,value = object_n)
    ws.cell(row = 1,column = (i+1)*2+1, value = '重量')



for key in people:
    ws.cell(row = row_n,column = 1,value = key)
    # print(people[key])
    for key_object in people[key]:
        if key_object == 'total':
            total_n = people[key][key_object]
        else:    
            if key_object in goods_name: 
                i = goods_name.index(key_object)
                ws.cell(row = row_n,column = (i+1)*2 , value = people[key][key_object][0])
                ws.cell(row = row_n,column = (i+1)*2+1 , value = people[key][key_object][1])
    ws.cell(row = row_n,column = (len(goods_name)+1)*2,value = total_n)
    row_n += 1
wb.save('./result/result.xlsx')
